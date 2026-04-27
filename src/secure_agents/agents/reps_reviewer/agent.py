"""Reps & warranties reviewer agent.

For each sorted contract, evaluates it against every representation and
warranty in the purchase agreement and produces an Excel matrix showing
which contracts trigger which reps and quoting the relevant language.

This is a **single-run** agent: scans the sorted category folders once,
processes all contracts in parallel, writes ``reps_analysis.xlsx``, and
stops.  It does not poll and does not use the job queue.

Configuration::

    agents:
      reps_reviewer:
        enabled: true
        reps_file: ./data/sample_reps.json  # JSON list of {id, title, text}
        output_root: /path/to/sorted/docs   # must match doc_sorter.output_root
        output_file: reps_analysis.xlsx     # written into output_root
        review_workers: 2                   # parallel LLM calls (one per contract)
        max_chars: 4000                     # contract text sent to the LLM per call
        tools:
          - text_extractor
          - file_manager
"""

from __future__ import annotations

import json
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import structlog

from secure_agents.core.base_agent import BaseAgent
from secure_agents.core.message_builder import MessageBuilder
from secure_agents.core.registry import register_agent
from secure_agents.core.schemas import validate_schema

from .prompts import REPS_ANALYSIS_SCHEMA, SYSTEM_PROMPT, build_review_instruction

logger = structlog.get_logger()

# Matches \u followed by fewer than 4 hex digits or non-hex chars — produced
# occasionally by smaller models when quoting contract language that contains
# backslashes.  Replace with a safe space to allow JSON parsing to succeed.
_INVALID_UNICODE_ESCAPE = re.compile(r'\\u(?![0-9a-fA-F]{4})')


def _sanitize_json(text: str) -> str:
    """Remove invalid \\uXXXX sequences that break json.loads."""
    return _INVALID_UNICODE_ESCAPE.sub(' ', text)


_SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".doc", ".pptx", ".xlsx"}

# Category folder names — must match CATEGORY_FOLDERS in doc_sorter/prompts.py
_CATEGORY_FOLDERS: dict[str, str] = {
    "nda":            "Non-disclosure agreements",
    "msa_company":    "MSAs (on company paper)",
    "msa_thirdparty": "MSAs (on third party paper)",
}

# Characters of contract text sent to the LLM per call.  Enough for the
# opening clauses, parties block, and key provisions on most legal documents.
_DEFAULT_MAX_CHARS = 4_000


@register_agent("reps_reviewer")
class RepsReviewerAgent(BaseAgent):
    """Evaluates every sorted contract against reps & warranties."""

    name = "reps_reviewer"
    description = "Review sorted contracts against reps & warranties from a purchase agreement"
    version = "0.1.0"
    features = [
        "Evaluates every sorted contract against every rep",
        "Quotes verbatim contract language for triggered reps",
        "Parallel processing — one LLM call per contract covers all reps",
        "Excel output: rows = contracts, columns = reps, cells = triggered/not",
        "Summary sheet with triggered counts per rep",
    ]

    def __init__(self, tools, provider, config=None, **kwargs):
        super().__init__(tools, provider, config, **kwargs)
        self.output_root: str = self.config.get("output_root", "./ai_generated")
        self.reps_file: str = self.config.get("reps_file", "")
        self.output_file: str = self.config.get("output_file", "reps_analysis.xlsx")
        self.max_chars: int = int(self.config.get("max_chars", _DEFAULT_MAX_CHARS))

    # ── Main loop (single-run) ────────────────────────────────────────────────

    def tick(self) -> None:
        # 1. Load reps
        if not self.reps_file:
            logger.error(
                "reps_reviewer.no_reps_file",
                msg="Set 'reps_file' in agent config to a JSON file of reps & warranties",
            )
            self._stop_event.set()
            return

        reps = self._load_reps()
        if not reps:
            self._stop_event.set()
            return

        logger.info("reps_reviewer.reps_loaded", count=len(reps))

        # 2. Discover contracts in sorted category folders
        contracts = self._scan_contracts()
        if not contracts:
            logger.warning(
                "reps_reviewer.no_contracts",
                output_root=self.output_root,
                msg="No contract files found in category folders. Run doc_sorter first.",
            )
            self._stop_event.set()
            return

        logger.info("reps_reviewer.contracts_found", count=len(contracts))

        # 3. Evaluate each contract against all reps (one LLM call per contract)
        workers = int(self.config.get("review_workers", 2))
        results: list[dict] = []

        logger.info(
            "reps_reviewer.reviewing_parallel",
            total=len(contracts),
            workers=workers,
            reps=len(reps),
        )

        with ThreadPoolExecutor(max_workers=workers) as executor:
            futures = {
                executor.submit(self._evaluate_contract, c, reps): c
                for c in contracts
            }
            for future in as_completed(futures):
                try:
                    result = future.result()
                    if result:
                        results.append(result)
                except Exception:
                    logger.exception("reps_reviewer.worker_error")

        logger.info(
            "reps_reviewer.review_complete",
            evaluated=len(results),
            total=len(contracts),
        )

        # 4. Write Excel output
        if results:
            self._write_excel(reps, results)

        # Single-run
        self._stop_event.set()

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _load_reps(self) -> list[dict]:
        """Load reps from the configured JSON file."""
        path = Path(self.reps_file).expanduser().resolve()
        if not path.is_file():
            logger.error("reps_reviewer.reps_file_not_found", path=str(path))
            return []
        try:
            reps = json.loads(path.read_text())
        except Exception as e:
            logger.error("reps_reviewer.reps_load_error", error=str(e))
            return []

        if not isinstance(reps, list):
            logger.error(
                "reps_reviewer.invalid_reps_format",
                msg="reps_file must contain a JSON array of {id, title, text} objects",
            )
            return []

        for rep in reps:
            if not all(k in rep for k in ("id", "title", "text")):
                logger.error(
                    "reps_reviewer.invalid_rep",
                    rep=rep,
                    msg="Each rep must have 'id', 'title', and 'text' fields",
                )
                return []

        return reps

    def _scan_contracts(self) -> list[dict]:
        """Scan sorted category folders and return file descriptors."""
        output_root = Path(self.output_root).expanduser().resolve()
        contracts: list[dict] = []

        for cat_key, folder_name in _CATEGORY_FOLDERS.items():
            folder = output_root / folder_name
            if not folder.is_dir():
                logger.debug("reps_reviewer.folder_missing", folder=str(folder))
                continue
            for f in sorted(folder.iterdir()):
                if f.is_file() and f.suffix.lower() in _SUPPORTED_EXTENSIONS:
                    contracts.append({
                        "path": str(f),
                        "filename": f.name,
                        "category": cat_key,
                        "category_name": folder_name,
                    })

        return contracts

    def _evaluate_contract(
        self, contract: dict, reps: list[dict]
    ) -> dict | None:
        """Run one LLM call evaluating a contract against all reps.

        Returns a result dict or None on failure.  If the first call produces
        invalid JSON (e.g. output was truncated by a token cap), we retry once
        with half the contract text so the response stays within limits.
        """
        text_ext = self.get_tool("text_extractor")

        extract = text_ext.execute(file_path=contract["path"])
        if "error" in extract:
            logger.warning(
                "reps_reviewer.extract_failed",
                filename=contract["filename"],
                error=extract["error"],
            )
            return None

        full_text = extract["text"]
        if not full_text.strip():
            logger.warning("reps_reviewer.empty_text", filename=contract["filename"])
            return None

        # Attempt with progressively shorter text: full max_chars, then half.
        for attempt, char_limit in enumerate([self.max_chars, self.max_chars // 2], 1):
            text = full_text[:char_limit]
            messages = (
                MessageBuilder(SYSTEM_PROMPT)
                .add_instruction(build_review_instruction(reps))
                .add_untrusted("contract", text)
                .build()
            )

            try:
                response = self.provider.complete(
                    messages, response_schema=REPS_ANALYSIS_SCHEMA
                )
            except Exception:
                logger.exception(
                    "reps_reviewer.llm_error",
                    filename=contract["filename"],
                    attempt=attempt,
                )
                if attempt == 1:
                    continue  # retry with shorter text
                return None

            ok, parsed = validate_schema(response.content, REPS_ANALYSIS_SCHEMA)
            if not ok:
                # Repair common model output quirks before giving up:
                # 1. Confidence values expressed as 0-100 instead of 0.0-1.0
                # 2. Invalid \uXXXX unicode escapes (model hallucinates hex digits)
                try:
                    cleaned = _sanitize_json(response.content)
                    raw = json.loads(cleaned)
                    for item in raw.get("results", []):
                        c = item.get("confidence")
                        if isinstance(c, (int, float)) and c > 1.0:
                            item["confidence"] = c / 100.0
                    ok, parsed = validate_schema(json.dumps(raw), REPS_ANALYSIS_SCHEMA)
                except Exception:
                    pass

            if ok:
                if attempt > 1:
                    logger.info(
                        "reps_reviewer.retry_succeeded",
                        filename=contract["filename"],
                        char_limit=char_limit,
                    )
                break

            logger.warning(
                "reps_reviewer.schema_invalid",
                filename=contract["filename"],
                attempt=attempt,
                error=parsed,
            )
            if attempt == 1:
                continue  # retry with shorter text
            return None

        # Index results by rep_id
        rep_results = {r["rep_id"]: r for r in parsed["results"]}
        triggered_count = sum(
            1 for r in parsed["results"] if r.get("triggered")
        )

        logger.info(
            "reps_reviewer.contract_evaluated",
            filename=contract["filename"],
            triggered=triggered_count,
            total_reps=len(reps),
        )

        return {
            "filename": contract["filename"],
            "category": contract["category"],
            "category_name": contract["category_name"],
            "rep_results": rep_results,
        }

    def _write_excel(self, reps: list[dict], results: list[dict]) -> None:
        """Write reps_analysis.xlsx to output_root."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error(
                "reps_reviewer.openpyxl_missing",
                msg="openpyxl is required. Run: pip install openpyxl",
            )
            return

        wb = openpyxl.Workbook()

        # ── Sheet 1: Full analysis matrix ─────────────────────────────────────
        ws = wb.active
        ws.title = "Analysis"

        # Styles
        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="1F4E79")   # dark blue
        trig_fill = PatternFill("solid", fgColor="FCE4D6")   # light orange
        ok_fill   = PatternFill("solid", fgColor="E2EFDA")   # light green
        na_fill   = PatternFill("solid", fgColor="F2F2F2")   # light grey
        wrap_top  = Alignment(wrap_text=True, vertical="top")

        # Header row: Contract | Category | [rep titles...]
        headers = ["Contract", "Category"] + [r["title"] for r in reps]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = wrap_top

        # Data rows sorted by category then filename
        sorted_results = sorted(
            results, key=lambda x: (x["category"], x["filename"])
        )
        for row_idx, res in enumerate(sorted_results, 2):
            ws.cell(row=row_idx, column=1, value=res["filename"]).alignment = wrap_top
            ws.cell(row=row_idx, column=2, value=res["category_name"]).alignment = wrap_top

            for col_idx, rep in enumerate(reps, 3):
                rep_res = res["rep_results"].get(rep["id"])
                if rep_res is None:
                    cell_val = "N/A"
                    fill = na_fill
                elif rep_res.get("triggered"):
                    conf    = rep_res.get("confidence", 0.0)
                    quoted  = (rep_res.get("quoted_language") or "").strip()
                    reason  = (rep_res.get("reasoning") or "").strip()
                    cell_val = f"TRIGGERED  (conf: {conf:.0%})"
                    if quoted:
                        cell_val += f'\n\n"{quoted}"'
                    if reason:
                        cell_val += f"\n\n{reason}"
                    fill = trig_fill
                else:
                    cell_val = "Not triggered"
                    fill = ok_fill

                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val)
                cell.alignment = wrap_top
                cell.fill = fill

        # Column widths
        ws.column_dimensions[get_column_letter(1)].width = 40   # Contract
        ws.column_dimensions[get_column_letter(2)].width = 22   # Category
        for i in range(len(reps)):
            ws.column_dimensions[get_column_letter(3 + i)].width = 45

        # Row heights (tall rows for quoted language)
        ws.row_dimensions[1].height = 30
        for row in ws.iter_rows(min_row=2, max_row=len(sorted_results) + 1):
            ws.row_dimensions[row[0].row].height = 90

        # Freeze header row + first two columns
        ws.freeze_panes = ws["C2"]

        # ── Sheet 2: Summary ──────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        s_hdr = ["Rep ID", "Title", "Triggered (# contracts)", "% of contracts reviewed"]
        for col, h in enumerate(s_hdr, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        total = len(results)
        for i, rep in enumerate(reps, 2):
            count = sum(
                1 for r in results
                if r["rep_results"].get(rep["id"], {}).get("triggered", False)
            )
            pct = f"{count / total:.0%}" if total else "0%"
            ws2.append([rep["id"], rep["title"], count, pct])

        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 45
        ws2.column_dimensions["C"].width = 28
        ws2.column_dimensions["D"].width = 28

        # ── Save ──────────────────────────────────────────────────────────────
        output_path = (
            Path(self.output_root).expanduser().resolve() / self.output_file
        )
        try:
            wb.save(str(output_path))
            logger.info(
                "reps_reviewer.excel_written",
                path=str(output_path),
                contracts=len(results),
                reps=len(reps),
            )
        except Exception as e:
            logger.error("reps_reviewer.excel_save_failed", error=str(e))
